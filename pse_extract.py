import io
import re
import zipfile
import xml.etree.ElementTree as ET
from typing import Callable, Optional

from pse_common import INDEXABLE_EXTS


RTF_DESTINATIONS = {
    "colortbl",
    "datastore",
    "fonttbl",
    "generator",
    "info",
    "listoverride",
    "listtable",
    "pict",
    "stylesheet",
    "themedata",
}

RTF_REPLACEMENTS = {
    "bullet": "*",
    "emdash": "-",
    "endash": "-",
    "lquote": "'",
    "rquote": "'",
    "ldblquote": '"',
    "rdblquote": '"',
    "line": "\n",
    "par": "\n",
    "tab": "\t",
}

PARTIAL_BYTE_CAP_EXTS = {
    ".txt", ".md", ".py", ".js", ".ts", ".json", ".yaml", ".yml", ".csv", ".log",
    ".html", ".htm", ".tex", ".rtf",
}


class TextBuilder:
    def __init__(self, char_cap: int) -> None:
        self.char_cap = max(0, int(char_cap))
        self.parts: list[str] = []
        self.total = 0

    @property
    def full(self) -> bool:
        return self.total >= self.char_cap

    def append(self, text: Optional[str]) -> None:
        if not text or self.full:
            return
        remaining = self.char_cap - self.total
        if remaining <= 0:
            return
        if len(text) > remaining:
            text = text[:remaining]
        self.parts.append(text)
        self.total += len(text)

    def build(self) -> str:
        return "".join(self.parts)


def supports_partial_byte_cap(ext: str) -> bool:
    return ext.lower() in PARTIAL_BYTE_CAP_EXTS


def extract_text(
    path: str,
    ext: str,
    char_cap: int,
    max_bytes: Optional[int] = None,
) -> Optional[str]:
    """
    Shared content extraction used by both indexing and snippets.

    Handler summary:
    - pdf: PyMuPDF if installed
    - html/htm: visible text only
    - tex/rtf: best-effort readable text
    - docx/odt/pptx: zipped XML text
    - xlsx: sheet values via openpyxl
    - doc/xls/ppt: Windows Office COM automation
    - everything else: plain text decode fallback
    """
    ext = ext.lower()
    if ext not in INDEXABLE_EXTS:
        return None

    if ext == ".pdf":
        return _extract_pdf_text(path, char_cap)
    if ext in {".html", ".htm"}:
        return _extract_html_text(path, char_cap, max_bytes=max_bytes)
    if ext == ".tex":
        return _extract_tex_text(path, char_cap, max_bytes=max_bytes)
    if ext == ".rtf":
        return _extract_rtf_text(path, char_cap, max_bytes=max_bytes)
    if ext == ".docx":
        return _extract_docx_text(path, char_cap)
    if ext == ".odt":
        return _extract_odt_text(path, char_cap)
    if ext == ".pptx":
        return _extract_pptx_text(path, char_cap)
    if ext == ".xlsx":
        return _extract_xlsx_text(path, char_cap)
    if ext == ".doc":
        return _extract_doc_via_com(path, char_cap)
    if ext == ".xls":
        return _extract_xls_via_com(path, char_cap)
    if ext == ".ppt":
        return _extract_ppt_via_com(path, char_cap)

    return _extract_plain_text(path, char_cap, max_bytes=max_bytes)


def _byte_cap(char_cap: int, multiplier: int = 4) -> int:
    return max(4096, min(max(1, int(char_cap)) * multiplier, 16_000_000))


def _effective_byte_cap(char_cap: int, max_bytes: Optional[int], multiplier: int = 4) -> int:
    cap = _byte_cap(char_cap, multiplier=multiplier)
    if max_bytes is None:
        return cap
    return max(1, min(cap, int(max_bytes)))


def _read_bytes(path: str, byte_cap: int) -> Optional[bytes]:
    try:
        with open(path, "rb") as fh:
            return fh.read(byte_cap)
    except (PermissionError, FileNotFoundError, OSError):
        return None


def _read_zip_member(path: str, member_name: str) -> Optional[bytes]:
    try:
        with zipfile.ZipFile(path) as archive:
            return archive.read(member_name)
    except (KeyError, FileNotFoundError, OSError, zipfile.BadZipFile):
        return None


def _decode_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _clean_visible_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\f\v]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[1]
    return tag


def _attr_local(elem: ET.Element, name: str, default: str = "") -> str:
    for key, value in elem.attrib.items():
        if _local_name(key) == name:
            return value
    return default


def _extract_plain_text(path: str, char_cap: int, max_bytes: Optional[int] = None) -> Optional[str]:
    data = _read_bytes(path, _effective_byte_cap(char_cap, max_bytes, multiplier=1))
    if data is None:
        return None
    return _decode_bytes(data)[:char_cap]


def _extract_pdf_text(path: str, char_cap: int) -> Optional[str]:
    # Lazy import keeps PDF support optional for the rest of the app.
    try:
        import fitz
    except ImportError:
        return None

    try:
        doc = fitz.open(path)
    except Exception:
        return None

    builder = TextBuilder(char_cap)
    try:
        for page in doc:
            text = page.get_text("text")
            if text:
                builder.append(text)
            if builder.full:
                break
    finally:
        doc.close()

    return _clean_visible_text(builder.build())


def _extract_html_text(path: str, char_cap: int, max_bytes: Optional[int] = None) -> Optional[str]:
    data = _read_bytes(path, _effective_byte_cap(char_cap, max_bytes, multiplier=4))
    if data is None:
        return None

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", _decode_bytes(data))
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        return _clean_visible_text(text)[:char_cap]

    soup = BeautifulSoup(_decode_bytes(data), "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return _clean_visible_text(soup.get_text("\n"))[:char_cap]


def _extract_tex_text(path: str, char_cap: int, max_bytes: Optional[int] = None) -> Optional[str]:
    data = _read_bytes(path, _effective_byte_cap(char_cap, max_bytes, multiplier=4))
    if data is None:
        return None

    text = _decode_bytes(data)
    text = "\n".join(re.sub(r"(?<!\\)%.*$", "", line) for line in text.splitlines())
    text = text.replace("\\\\", "\n")
    text = re.sub(r"\\(par|newline|linebreak|item)\b", "\n", text)
    text = re.sub(r"\\(begin|end)\{[^{}]+\}", "\n", text)

    for _ in range(4):
        updated = re.sub(r"\\[a-zA-Z@]+\*?(?:\[[^\]]*\])?\{([^{}]*)\}", r"\1", text)
        if updated == text:
            break
        text = updated

    text = re.sub(r"\\[a-zA-Z@]+\*?(?:\[[^\]]*\])?", " ", text)
    text = text.replace("{", " ").replace("}", " ").replace("~", " ")
    return _clean_visible_text(text)[:char_cap]


def _rtf_to_text(content: str, char_cap: int) -> str:
    stack: list[tuple[int, bool]] = []
    ignorable = False
    ucskip = 1
    curskip = 0
    builder = TextBuilder(char_cap)
    i = 0

    while i < len(content) and not builder.full:
        ch = content[i]

        if ch == "{":
            stack.append((ucskip, ignorable))
        elif ch == "}":
            if stack:
                ucskip, ignorable = stack.pop()
            curskip = 0
        elif ch == "\\":
            i += 1
            if i >= len(content):
                break
            ch = content[i]

            if ch in "{}\\":
                if not ignorable:
                    builder.append(ch)
                curskip = 0
            elif ch == "'":
                hex_value = content[i + 1:i + 3]
                if len(hex_value) == 2:
                    try:
                        decoded = bytes.fromhex(hex_value).decode("cp1252")
                    except ValueError:
                        decoded = ""
                    if not ignorable:
                        builder.append(decoded)
                    i += 2
                curskip = 0
            elif ch in "~_-":
                if not ignorable:
                    builder.append(" " if ch == "~" else "-")
                curskip = 0
            elif ch == "*":
                ignorable = True
                curskip = 0
            elif ch.isalpha():
                match = re.match(r"([a-zA-Z]+)(-?\d+)? ?", content[i:])
                if not match:
                    curskip = 0
                else:
                    word = match.group(1)
                    arg = match.group(2)
                    i += len(match.group(0)) - 1

                    if word in RTF_DESTINATIONS:
                        ignorable = True
                    elif word == "uc":
                        ucskip = int(arg or 1)
                    elif word == "u":
                        if not ignorable:
                            codepoint = int(arg or 0)
                            if codepoint < 0:
                                codepoint += 65536
                            builder.append(chr(codepoint))
                        curskip = ucskip
                    else:
                        replacement = RTF_REPLACEMENTS.get(word)
                        if replacement and not ignorable:
                            builder.append(replacement)
                        curskip = 0
            else:
                curskip = 0
        else:
            if curskip > 0:
                curskip -= 1
            elif not ignorable:
                builder.append(ch)

        i += 1

    return _clean_visible_text(builder.build())


def _extract_rtf_text(path: str, char_cap: int, max_bytes: Optional[int] = None) -> Optional[str]:
    data = _read_bytes(path, _effective_byte_cap(char_cap, max_bytes, multiplier=4))
    if data is None:
        return None
    return _rtf_to_text(_decode_bytes(data), char_cap)


def _append_openxml_text(
    builder: TextBuilder,
    xml_bytes: bytes,
    text_tags: set[str],
    break_tags: set[str],
    block_tags: set[str],
) -> None:
    try:
        events = ET.iterparse(io.BytesIO(xml_bytes), events=("start", "end"))
        for event, elem in events:
            name = _local_name(elem.tag)
            if event == "start":
                if name in break_tags:
                    builder.append("\n")
            else:
                if name in text_tags and elem.text:
                    builder.append(elem.text)
                elif name in block_tags:
                    builder.append("\n")
                elem.clear()
            if builder.full:
                break
    except ET.ParseError:
        return


def _extract_docx_text(path: str, char_cap: int) -> Optional[str]:
    xml_bytes = _read_zip_member(path, "word/document.xml")
    if xml_bytes is None:
        return None

    builder = TextBuilder(char_cap)
    _append_openxml_text(builder, xml_bytes, {"t"}, {"br", "cr", "tab"}, {"p"})
    return _clean_visible_text(builder.build())


def _extract_pptx_text(path: str, char_cap: int) -> Optional[str]:
    try:
        with zipfile.ZipFile(path) as archive:
            slide_parts = sorted(
                name for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            )
            builder = TextBuilder(char_cap)
            for slide_name in slide_parts:
                _append_openxml_text(
                    builder,
                    archive.read(slide_name),
                    {"t"},
                    {"br"},
                    {"p"},
                )
                builder.append("\n")
                if builder.full:
                    break
    except (FileNotFoundError, OSError, zipfile.BadZipFile):
        return None

    return _clean_visible_text(builder.build())


def _append_odt_node(builder: TextBuilder, elem: ET.Element) -> None:
    if builder.full:
        return

    name = _local_name(elem.tag)
    if name == "s":
        count = int(_attr_local(elem, "c", "1") or "1")
        builder.append(" " * max(1, count))
    elif name == "tab":
        builder.append("\t")
    elif name == "line-break":
        builder.append("\n")

    if elem.text and name != "s":
        builder.append(elem.text)

    for child in list(elem):
        _append_odt_node(builder, child)
        if builder.full:
            return
        if child.tail:
            builder.append(child.tail)

    if name in {"p", "h", "list-item"}:
        builder.append("\n")


def _extract_odt_text(path: str, char_cap: int) -> Optional[str]:
    xml_bytes = _read_zip_member(path, "content.xml")
    if xml_bytes is None:
        return None

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return None

    builder = TextBuilder(char_cap)
    _append_odt_node(builder, root)
    return _clean_visible_text(builder.build())


def _extract_xlsx_text(path: str, char_cap: int) -> Optional[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    builder = TextBuilder(char_cap)
    try:
        for sheet in workbook.worksheets:
            builder.append(sheet.title)
            builder.append("\n")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value not in (None, "")]
                if values:
                    builder.append(" ".join(values))
                    builder.append("\n")
                if builder.full:
                    break
            if builder.full:
                break
    finally:
        workbook.close()

    return _clean_visible_text(builder.build())


def _import_pywin32():
    # Legacy Office extraction is best-effort and depends on local installs.
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return None, None
    return pythoncom, win32com.client


def _safe_invoke(obj, method_name: str, *args) -> None:
    if obj is None:
        return
    try:
        getattr(obj, method_name)(*args)
    except Exception:
        pass


def _run_with_com_app(
    prog_id: str,
    worker: Callable[[object], Optional[str]],
    cleanup: Optional[Callable[[], None]] = None,
) -> Optional[str]:
    pythoncom, win32_client = _import_pywin32()
    if pythoncom is None or win32_client is None:
        return None

    app = None
    pythoncom.CoInitialize()
    try:
        app = win32_client.DispatchEx(prog_id)
        return worker(app)
    except Exception:
        return None
    finally:
        if cleanup is not None:
            cleanup()
        _safe_invoke(app, "Quit")
        pythoncom.CoUninitialize()


def _extract_doc_via_com(path: str, char_cap: int) -> Optional[str]:
    document = None

    def cleanup() -> None:
        _safe_invoke(document, "Close", False)

    def worker(app: object) -> Optional[str]:
        nonlocal document
        app.Visible = False
        app.DisplayAlerts = 0
        document = app.Documents.Open(
            path,
            ReadOnly=True,
            AddToRecentFiles=False,
            Visible=False,
            ConfirmConversions=False,
            NoEncodingDialog=True,
        )
        return _clean_visible_text((document.Range().Text or "")[:char_cap])

    return _run_with_com_app("Word.Application", worker, cleanup=cleanup)


def _iter_excel_rows(values):
    if values in (None, ""):
        return
    if not isinstance(values, (list, tuple)):
        yield (values,)
        return
    if values and not isinstance(values[0], (list, tuple)):
        yield tuple(values)
        return
    for row in values:
        if isinstance(row, (list, tuple)):
            yield tuple(row)
        else:
            yield (row,)


def _extract_xls_via_com(path: str, char_cap: int) -> Optional[str]:
    workbook = None

    def cleanup() -> None:
        _safe_invoke(workbook, "Close", False)

    def worker(app: object) -> Optional[str]:
        nonlocal workbook
        app.Visible = False
        app.DisplayAlerts = False
        workbook = app.Workbooks.Open(path, ReadOnly=True, AddToMru=False, Notify=False)

        builder = TextBuilder(char_cap)
        for sheet in workbook.Worksheets:
            builder.append(getattr(sheet, "Name", ""))
            builder.append("\n")
            used_range = getattr(sheet, "UsedRange", None)
            values = getattr(used_range, "Value", None)
            for row in _iter_excel_rows(values):
                rendered = [str(value) for value in row if value not in (None, "")]
                if rendered:
                    builder.append(" ".join(rendered))
                    builder.append("\n")
                if builder.full:
                    break
            if builder.full:
                break
        return _clean_visible_text(builder.build())

    return _run_with_com_app("Excel.Application", worker, cleanup=cleanup)


def _shape_text(shape) -> str:
    try:
        if getattr(shape, "HasTextFrame", 0):
            text_frame = shape.TextFrame
            if getattr(text_frame, "HasText", True):
                return getattr(text_frame.TextRange, "Text", "") or ""
    except Exception:
        pass
    try:
        text_frame2 = shape.TextFrame2
        if getattr(text_frame2, "HasText", False):
            return getattr(text_frame2.TextRange, "Text", "") or ""
    except Exception:
        pass
    return ""


def _extract_ppt_via_com(path: str, char_cap: int) -> Optional[str]:
    presentation = None

    def cleanup() -> None:
        _safe_invoke(presentation, "Close")

    def worker(app: object) -> Optional[str]:
        nonlocal presentation
        presentation = app.Presentations.Open(path, ReadOnly=True, WithWindow=False)

        builder = TextBuilder(char_cap)
        for slide in presentation.Slides:
            for shape in slide.Shapes:
                text = _shape_text(shape)
                if text:
                    builder.append(text)
                    builder.append("\n")
                if builder.full:
                    break
            if builder.full:
                break
        return _clean_visible_text(builder.build())

    return _run_with_com_app("PowerPoint.Application", worker, cleanup=cleanup)

